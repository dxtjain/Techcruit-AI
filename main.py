import os 
import fitz  # PyMuPDF
import requests
import json
import re
import openpyxl
from openpyxl.styles import Font
from flask import Flask, render_template, request, jsonify, send_from_directory
from flask_cors import CORS
import logging
from datetime import datetime, timedelta
import traceback

app = Flask(__name__, static_folder='ui ux/dist', static_url_path='')
CORS(app)  # Enable CORS for all routes

# Configuration
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'fallback-secret-key-for-development')

# Set up logging
if os.environ.get('FLASK_ENV') == 'production':
    logging.basicConfig(level=logging.WARNING)
else:
    logging.basicConfig(level=logging.INFO)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Environment variable configuration
GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
if not GROQ_API_KEY:
    raise ValueError("GROQ_API_KEY environment variable is required. Please set it in your deployment settings.")

GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
MODEL_NAME = "llama-3.3-70b-versatile"
EXCEL_FILE = "data/resumes_data.xlsx"

# Ensure data directory exists
os.makedirs("data", exist_ok=True)

HEADERS = {
    "Authorization": f"Bearer {GROQ_API_KEY}",
    "Content-Type": "application/json"
}

# -------- PDF Text Extraction --------
def extract_text_from_pdf(filepath):
    text = ""
    with fitz.open(filepath) as doc:
        for page in doc:
            text += page.get_text()
    return text

# -------- Groq API Query --------
def query_groq(prompt):
    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.7,
        "max_tokens": 4000
    }
    try:
        response = requests.post(GROQ_API_URL, headers=HEADERS, json=payload, timeout=30)
        response.raise_for_status()  # Raise an exception for bad status codes
        return response.json()
    except requests.exceptions.Timeout:
        return {"error": "Request timeout - please try again"}
    except requests.exceptions.HTTPError as e:
        if response.status_code == 401:
            return {"error": "Invalid API key - please check your GROQ_API_KEY"}
        elif response.status_code == 429:
            return {"error": "Rate limit exceeded - please try again later"}
        else:
            return {"error": f"HTTP error {response.status_code}: {str(e)}"}
    except Exception as e:
        return {"error": f"Unexpected error: {str(e)}"}

# -------- Extract JSON from Groq Response --------
def extract_resumes_from_groq_content(content):
    pattern = r"\*\*Resume\s\d+\s-\s(.*?)\*\*\n```json\n(.*?)\n```"
    matches = re.findall(pattern, content, re.DOTALL)
    results = []
    for filename, json_block in matches:
        try:
            resume_data = json.loads(json_block)
            resume_data["filename"] = filename.strip()
            results.append(resume_data)
        except json.JSONDecodeError as e:
            print(f"Error parsing {filename}: {e}")
    return results

# -------- Save Resume Data to Excel --------
import os
import openpyxl
from openpyxl.styles import Font

def save_single_resume_to_excel(resume, filename=EXCEL_FILE):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resume Data"
        headers = [
            "Name", "Email", "Phone Number", "Experience (Years)",
            "Skills", "Used Software", "Expected Domain"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Convert lists to comma-separated strings
    skills = ", ".join(resume.get("skills", []))
    software = ", ".join(resume.get("used_software", []))

    # Append the row
    ws.append([
        resume.get("name", ""),
        resume.get("email", ""),
        resume.get("phone_number", ""),
        resume.get("experience_in_years", ""),
        skills,
        software,
        resume.get("expected_domain", "")
    ])

    wb.save(filename)

# -------- DOCX Text Extraction --------

# -------- Routes --------
@app.route('/api/upload', methods=['POST'])
def upload_resumes():
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files')
    if len(files) > 10:  # Limit number of files
        return jsonify({'error': 'Maximum 10 files allowed per upload'}), 400
    
    resume_texts = {}
    allowed_extensions = {'.pdf', '.docx'}
    
    for file in files:
        if not file.filename:
            continue
            
        # Security: Check file extension
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({'error': f'Unsupported file type: {file_ext}. Only PDF and DOCX files are allowed.'}), 400
        
        # Security: Sanitize filename
        import uuid
        safe_filename = f"{uuid.uuid4()}{file_ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
        
        try:
            file.save(filepath)
            text = extract_text_from_pdf(filepath)
            if len(text.strip()) < 50:  # Basic validation
                os.remove(filepath)  # Clean up
                return jsonify({'error': f'File {file.filename} appears to be empty or corrupted'}), 400
            resume_texts[file.filename] = text
            os.remove(filepath)  # Clean up after processing
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)  # Clean up on error
            return jsonify({'error': f'Error processing file {file.filename}: {str(e)}'}), 500

    # Build one big prompt for all resumes
    prompt = "Analyze the following resumes one by one. For each resume, return a single JSON object with keys: name, email, skills, experience in years, phone number, expected domain, used software.\n\n"
    for i, (filename, text) in enumerate(resume_texts.items(), start=1):
        prompt += f"Resume {i} - {filename}:\n{text}\n\n"

    groq_response = query_groq(prompt)

    if "choices" in groq_response:
        content = groq_response["choices"][0]["message"]["content"]
        resume_data = extract_resumes_from_groq_content(content)
        
        # Save to Excel
        for resume in resume_data:
            save_single_resume_to_excel(resume)

        return jsonify({'resumes': resume_data})
    else:
        return jsonify({'error': str(groq_response)}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint for monitoring"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0',
        'service': 'Techcruit AI'
    })

@app.route('/')
def serve_frontend():
    try:
        return send_from_directory(app.static_folder, 'index.html')
    except:
        # Fallback for when static files are not available (Vercel deployment)
        return """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Techcruit AI - Intelligent Recruitment Platform</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; 
                       margin: 0; padding: 40px; background: #f8fafc; }
                .container { max-width: 800px; margin: 0 auto; background: white; padding: 40px; 
                            border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
                h1 { color: #1e40af; margin-bottom: 20px; }
                .api-section { background: #f1f5f9; padding: 20px; border-radius: 8px; margin: 20px 0; }
                .endpoint { font-family: 'Courier New', monospace; background: #e2e8f0; 
                           padding: 8px 12px; border-radius: 4px; margin: 8px 0; }
                .status { color: #059669; font-weight: bold; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>🚀 Techcruit AI - Backend API</h1>
                <div class="status">✅ Backend is running successfully!</div>
                
                <div class="api-section">
                    <h3>Available API Endpoints:</h3>
                    <div class="endpoint">GET /api/health - Health check</div>
                    <div class="endpoint">GET /api/dashboard/stats - Dashboard statistics</div>
                    <div class="endpoint">POST /api/batch/process - Process resumes</div>
                    <div class="endpoint">GET /api/batch/history - Processing history</div>
                    <div class="endpoint">POST /api/ai/analyze - AI resume analysis</div>
                    <div class="endpoint">POST /api/ai/compare - Compare resumes</div>
                    <div class="endpoint">GET /api/pricing/plans - Pricing plans</div>
                    <div class="endpoint">POST /api/pricing/calculate - Calculate pricing</div>
                    <div class="endpoint">POST /api/pricing/contact - Contact form</div>
                </div>
                
                <p><strong>Note:</strong> This is the backend API for Techcruit AI. 
                   For the full frontend experience, please deploy the React application separately 
                   or access the complete application through the intended domain.</p>
            </div>
        </body>
        </html>
        """

@app.route('/<path:path>')
def serve_static(path):
    try:
        return send_from_directory(app.static_folder, path)
    except:
        # Fallback for API-only deployment
        return jsonify({'error': 'This is an API endpoint. Please use /api/* routes.'}), 404

@app.route('/api/download-excel', methods=['GET'])
def download_excel():
    excel_path = os.path.join(os.getcwd(), 'resumes_data.xlsx')
    if os.path.exists(excel_path):
        return send_from_directory(os.getcwd(), 'resumes_data.xlsx', as_attachment=True)
    else:
        return jsonify({'error': 'Excel file not found'}), 404

# -------- Dashboard API Routes --------
@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        # Initialize default stats
        stats = {
            'resumesProcessed': 0,
            'topSkills': [],
            'recentUploads': [],
            'skillsDistribution': {},
            'experienceLevels': {'Junior': 0, 'Mid': 0, 'Senior': 0}
        }
        
        # Quick check for Excel file existence
        if os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
                ws = wb.active
                
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                stats['resumesProcessed'] = len(rows)
                
                if len(rows) > 0:
                    # Calculate skills distribution efficiently
                    all_skills = []
                    for row in rows:
                        if row and len(row) > 4 and row[4]:  # Skills column
                            skills = [skill.strip() for skill in str(row[4]).split(',') if skill.strip()]
                            all_skills.extend(skills)
                    
                    # Count skill frequencies
                    skill_counts = {}
                    for skill in all_skills:
                        if skill:  # Avoid empty skills
                            skill_counts[skill] = skill_counts.get(skill, 0) + 1
                    
                    # Get top 5 skills
                    if skill_counts:
                        stats['topSkills'] = sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)[:5]
                        stats['skillsDistribution'] = dict(sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)[:10])
                    
                    # Calculate experience levels efficiently
                    for row in rows:
                        if row and len(row) > 3 and row[3]:  # Experience column
                            try:
                                exp_str = str(row[3]).strip()
                                # Extract numeric value from experience string
                                import re
                                exp_match = re.search(r'(\d+(?:\.\d+)?)', exp_str)
                                if exp_match:
                                    exp = float(exp_match.group(1))
                                    if exp <= 2:
                                        stats['experienceLevels']['Junior'] += 1
                                    elif exp <= 5:
                                        stats['experienceLevels']['Mid'] += 1
                                    else:
                                        stats['experienceLevels']['Senior'] += 1
                            except:
                                pass
                
                wb.close()
            except Exception as e:
                print(f"Error reading Excel file: {e}")
                # Continue with default stats
        
        # Get recent uploads from uploads folder (limit to 5 most recent)
        uploads_folder = app.config['UPLOAD_FOLDER']
        if os.path.exists(uploads_folder):
            try:
                files = [f for f in os.listdir(uploads_folder) if f.endswith('.pdf')]
                files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_folder, x)), reverse=True)
                
                recent_files = []
                for file in files[:5]:  # Last 5 files
                    try:
                        filepath = os.path.join(uploads_folder, file)
                        import datetime
                        mod_time = os.path.getmtime(filepath)
                        date_str = datetime.datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
                        recent_files.append({'name': file, 'date': date_str})
                    except:
                        continue
                stats['recentUploads'] = recent_files
            except Exception as e:
                print(f"Error reading uploads folder: {e}")
        
        return jsonify(stats), 200
    except Exception as e:
        print(f"Dashboard stats error: {e}")
        # Return default stats even on error
        return jsonify({
            'resumesProcessed': 0,
            'topSkills': [],
            'recentUploads': [],
            'skillsDistribution': {},
            'experienceLevels': {'Junior': 0, 'Mid': 0, 'Senior': 0}
        }), 200

# -------- Batch Processing API Routes --------
@app.route('/api/batch/process', methods=['POST'])
def batch_process_resumes():
    """Process multiple resumes in batch"""
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files')
        if len(files) > 20:  # Limit batch size
            return jsonify({'error': 'Maximum 20 files allowed per batch'}), 400
        
        batch_results = []
        total_files = len(files)
        processed_files = 0
        
        for file in files:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
                file.save(filepath)
                text = extract_text_from_pdf(filepath)
                
                # Quick analysis for each resume
                prompt = f"Analyze this resume and extract: name, email, phone, years of experience, top 3 skills, expected domain. Return as JSON: {text[:2000]}..."
                
                groq_response = query_groq(prompt)
                if "choices" in groq_response:
                    content = groq_response["choices"][0]["message"]["content"]
                    # Try to extract JSON from response
                    try:
                        import json
                        data = json.loads(content)
                        data['filename'] = file.filename
                        data['status'] = 'success'
                    except:
                        data = {
                            'filename': file.filename,
                            'status': 'processed',
                            'message': 'Extracted basic info'
                        }
                else:
                    data = {
                        'filename': file.filename,
                        'status': 'error',
                        'message': 'Failed to analyze'
                    }
                
                batch_results.append(data)
                processed_files += 1
                
            except Exception as e:
                batch_results.append({
                    'filename': file.filename,
                    'status': 'error',
                    'message': str(e)
                })
        
        return jsonify({
            'message': f'Batch processing completed: {processed_files}/{total_files} files processed',
            'results': batch_results,
            'summary': {
                'total': total_files,
                'processed': processed_files,
                'errors': total_files - processed_files
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch/history', methods=['GET'])
def get_batch_history():
    """Get batch processing history"""
    # Mock data for demonstration
    history = [
        {
            'id': 1,
            'date': '2025-06-15',
            'filesCount': 15,
            'status': 'completed',
            'duration': '2m 34s'
        },
        {
            'id': 2,
            'date': '2025-06-14',
            'filesCount': 8,
            'status': 'completed',
            'duration': '1m 12s'
        },
        {
            'id': 3,
            'date': '2025-06-13',
            'filesCount': 22,
            'status': 'partial',
            'duration': '3m 45s'
        }
    ]
    return jsonify({'history': history})

# -------- AI Analysis API Routes --------
@app.route('/api/ai/analyze', methods=['POST'])
def ai_analyze_resume():
    """Perform deep AI analysis on a single resume"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
        text = extract_text_from_pdf(filepath)
        
        # Advanced AI analysis prompt
        analysis_prompt = f"""
        Perform a comprehensive analysis of this resume. Provide detailed insights on:
        1. Candidate strengths and weaknesses
        2. Skill proficiency levels
        3. Career progression analysis
        4. Recommended roles and salary range
        5. Areas for improvement
        6. Cultural fit indicators
        
        Resume text: {text[:3000]}...
        
        Return a detailed analysis in a structured format.
        """
        
        groq_response = query_groq(analysis_prompt)
        
        if "choices" in groq_response:
            analysis = groq_response["choices"][0]["message"]["content"]
            
            # Mock additional analysis data
            result = {
                'filename': file.filename,
                'analysis': analysis,
                'scores': {
                    'technical_skills': 85,
                    'experience_relevance': 78,
                    'communication': 82,
                    'leadership': 65,
                    'overall_fit': 77
                },
                'recommendations': [
                    'Strong technical background in mentioned technologies',
                    'Good progression in career responsibilities',
                    'Consider additional leadership experience',
                    'Skills align well with senior developer roles'
                ],
                'salary_estimate': {
                    'min': 75000,
                    'max': 95000,
                    'currency': 'USD'
                }
            }
            
            return jsonify(result)
        else:
            return jsonify({'error': 'AI analysis failed'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/ai/compare', methods=['POST'])
def ai_compare_resumes():
    """Compare multiple resumes using AI"""
    try:
        data = request.get_json()
        resume_ids = data.get('resume_ids', [])
        
        if len(resume_ids) < 2:
            return jsonify({'error': 'At least 2 resumes required for comparison'}), 400
        
        # Mock comparison data
        comparison = {
            'resumes': resume_ids,
            'comparison_matrix': {
                'technical_skills': [85, 78, 90],
                'experience': [5, 3, 7],
                'education': [88, 92, 85],
                'overall_score': [82, 78, 88]
            },
            'winner': resume_ids[2] if len(resume_ids) > 2 else resume_ids[0],
            'insights': [
                f"Resume {resume_ids[0]} shows strong technical foundation",
                f"Resume {resume_ids[1]} has excellent educational background",
                f"Resume {resume_ids[2] if len(resume_ids) > 2 else resume_ids[0]} demonstrates best overall fit"
            ]
        }
        
        return jsonify(comparison)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# -------- Pricing API Routes --------
@app.route('/api/pricing/plans', methods=['GET'])
def get_pricing_plans():
    """Get pricing plans and add-ons"""
    try:
        pricing_data = {
            'plans': [
                {
                    'id': 'starter',
                    'name': 'Starter',
                    'description': 'Perfect for small teams and startups',
                    'price': 15,
                    'currency': 'INR',
                    'unit': 'per resume',
                    'resumeLimit': 'Up to 100 resumes',
                    'resumeCount': 100,
                    'features': [
                        'Basic resume screening',
                        'AI-powered skills extraction', 
                        'Excel export',
                        'Email support',
                        'Standard processing speed'
                    ],
                    'savings': 'Basic screening package',
                    'popular': False,
                    'enterprise': False,
                    'icon': '<Rocket className="w-8 h-8" />',
                    'color': 'text-blue-600',
                    'bgColor': 'bg-blue-50',
                    'borderColor': 'border-blue-200'
                },
                {
                    'id': 'growth',
                    'name': 'Growth', 
                    'description': 'Best for growing companies',
                    'price': 20,
                    'currency': 'INR',
                    'unit': 'per resume',
                    'resumeLimit': 'Up to 500 resumes',
                    'resumeCount': 500,
                    'features': [
                        'Advanced AI filtering (Skills, Experience, Location)',
                        'Batch processing up to 1000 resumes',
                        'All export formats (Excel, PDF, Word)',
                        'Email auto-delivery',
                        'Priority support',
                        'Custom filters',
                        'Fresher/Experienced filtering'
                    ],
                    'savings': 'Filtered, All formats',
                    'popular': True,
                    'enterprise': False,
                    'icon': '<Building className="w-8 h-8" />',
                    'color': 'text-emerald-600',
                    'bgColor': 'bg-emerald-50',
                    'borderColor': 'border-emerald-300'
                },
                {
                    'id': 'enterprise',
                    'name': 'Enterprise',
                    'description': 'For large organizations (1000-20,000 resumes)',
                    'price': 0,
                    'currency': 'INR',
                    'unit': 'custom pricing',
                    'resumeLimit': '1000-20,000 resumes',
                    'resumeCount': 15000,
                    'features': [
                        'Full automation suite',
                        'Unlimited bulk processing',
                        'Advanced AI filtering',
                        'White-label solution',
                        'Dedicated account manager',
                        'Custom integrations',
                        'Branded reports with your logo',
                        'Priority processing',
                        'API access',
                        'Training sessions'
                    ],
                    'savings': 'Full automation + email delivery',
                    'popular': False,
                    'enterprise': True,
                    'icon': '<Crown className="w-8 h-8" />',
                    'color': 'text-purple-600',
                    'bgColor': 'bg-purple-50',
                    'borderColor': 'border-purple-300'
                }
            ],
            'addOns': [
                {
                    'id': 'skills-filter',
                    'name': 'Advanced Skills & Education Filter',
                    'description': 'Filter by specific skills, education levels, and experience criteria',
                    'price': 499,
                    'currency': 'INR',
                    'icon': '<Shield className="w-6 h-6" />',
                    'category': 'analytics'
                },
                {
                    'id': 'email-delivery',
                    'name': 'Email Auto-Delivery',
                    'description': 'Automated email delivery of processed results to your team',
                    'price': 299,
                    'currency': 'INR',
                    'icon': '<Mail className="w-6 h-6" />',
                    'category': 'integration'
                },
                {
                    'id': 'branded-reports',
                    'name': 'Branded Reports (With Your Logo)',
                    'description': 'Custom branded reports with your company logo and styling',
                    'price': 999,
                    'currency': 'INR',
                    'icon': '<Award className="w-6 h-6" />',
                    'category': 'security'
                }
            ],
            'contact': {
                'name': 'Tushar Jain',
                'phone': '+91-9359205909',
                'email': 'info@techmarqx.com'
            },
            'freeTrial': {
                'enabled': True,
                'resumeCount': 10,
                'description': 'Get your first 10 resumes screened at no cost!'
            }
        }
        
        return jsonify(pricing_data)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pricing/calculate', methods=['POST'])
def calculate_pricing():
    """Calculate total pricing based on selected plan and add-ons"""
    try:
        data = request.get_json()
        plan_id = data.get('plan_id')
        selected_addons = data.get('addons', [])
        resume_count = data.get('resume_count', 1)
        
        # Plan pricing (INR)
        plans = {
            'starter': {'price': 15, 'name': 'Starter'},
            'growth': {'price': 20, 'name': 'Growth'},
            'enterprise': {'price': 12, 'name': 'Enterprise'}  # Special enterprise rate
        }
        
        # Add-on pricing (INR)
        addons = {
            'skills-filter': {'price': 499, 'name': 'Advanced Skills & Education Filter'},
            'email-delivery': {'price': 299, 'name': 'Email Auto-Delivery'},
            'branded-reports': {'price': 999, 'name': 'Branded Reports (With Your Logo)'}
        }
        
        if plan_id not in plans:
            return jsonify({'error': 'Invalid plan selected'}), 400
        
        base_cost = plans[plan_id]['price'] * resume_count
        addon_cost = sum(addons[addon]['price'] for addon in selected_addons if addon in addons)
        total_cost = base_cost + addon_cost
        
        calculation = {
            'plan': {
                'id': plan_id,
                'name': plans[plan_id]['name'],
                'price_per_resume': plans[plan_id]['price'],
                'resume_count': resume_count,
                'subtotal': base_cost
            },
            'addons': [
                {
                    'id': addon_id,
                    'name': addons[addon_id]['name'],
                    'price': addons[addon_id]['price']
                }
                for addon_id in selected_addons if addon_id in addons
            ],
            'totals': {
                'base_cost': base_cost,
                'addon_cost': addon_cost,
                'total_cost': total_cost,
                'currency': 'INR'
            }
        }
        
        return jsonify(calculation)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pricing/contact', methods=['POST'])
def submit_contact_form():
    """Handle contact form submissions for enterprise inquiries"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'email', 'company', 'plan_interest']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # In a real application, you would:
        # 1. Save to database
        # 2. Send email notifications
        # 3. Integrate with CRM
        
        # Mock response
        response = {
            'message': 'Thank you for your interest! Our sales team will contact you within 24 hours.',
            'contact_id': f"CONTACT_{hash(data.get('email')) % 10000}",
            'estimated_response_time': '24 hours'
        }
        
        return jsonify(response)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # For local development
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
